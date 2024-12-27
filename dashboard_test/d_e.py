import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from io import BytesIO


class CampaignAnalyzer:
    def __init__(self):
        self.TAG_CRITERIA = {
            'HIGH_TRAFFIC': {'impressions': 5000, 'label': '流量高'},
            'LOW_ACOS': {'acos': 20, 'label': 'ACOS優'},
            'HIGH_CONVERSION': {'conversion': 15, 'label': '高轉換'},
            'POOR_PERFORMANCE': {
                'minSales': 10,
                'minClicks': 10,
                'maxAcos': 100,
                'label': '低成效'
            }
        }

    def load_data(self, file1_path, file2_path):
        df1 = pd.read_csv(file1_path)
        df2 = pd.read_csv(file2_path)
        return pd.concat([df1, df2])

    def clean_numeric_column(self, value):
        if pd.isna(value) or value == '':
            return 0
        return float(str(value).replace('$', '').replace('%', '').replace(',', ''))

    def process_data(self, df):
        # 清理數值欄位
        df['Sales'] = df['7 Day Total Sales '].apply(self.clean_numeric_column)
        df['Spend'] = df['Spend'].apply(self.clean_numeric_column)

        # 計算每個廣告活動的指標
        campaign_metrics = df.groupby('Campaign Name').agg({
            'Sales': 'sum',
            'Spend': 'sum',
            '7 Day Total Orders (#)': 'sum',
            'Impressions': 'sum',
            'Clicks': 'sum'
        }).reset_index()

        # 計算其他指標
        campaign_metrics['CTR'] = (
            campaign_metrics['Clicks'] / campaign_metrics['Impressions'] * 100).fillna(0)
        campaign_metrics['Conversion_Rate'] = (
            campaign_metrics['7 Day Total Orders (#)'] / campaign_metrics['Clicks'] * 100).fillna(0)
        # 修正 ACOS 計算方式
        campaign_metrics['ACOS'] = np.where(
            campaign_metrics['Sales'] > 0,
            (campaign_metrics['Spend'] / campaign_metrics['Sales'] * 100),
            0  # 當銷售額為 0 時，ACOS 設為 0
        )
        campaign_metrics['ROAS'] = np.where(
            campaign_metrics['Spend'] > 0,
            campaign_metrics['Sales'] / campaign_metrics['Spend'],
            0  # 當支出為 0 時，ROAS 設為 0
        )

        return campaign_metrics

    def get_campaign_tags(self, row):
        tags = []
        if row['Impressions'] >= self.TAG_CRITERIA['HIGH_TRAFFIC']['impressions']:
            tags.append(self.TAG_CRITERIA['HIGH_TRAFFIC']['label'])
        if 0 < row['ACOS'] < self.TAG_CRITERIA['LOW_ACOS']['acos']:
            tags.append(self.TAG_CRITERIA['LOW_ACOS']['label'])
        if row['Conversion_Rate'] >= self.TAG_CRITERIA['HIGH_CONVERSION']['conversion']:
            tags.append(self.TAG_CRITERIA['HIGH_CONVERSION']['label'])
        if (row['Sales'] < self.TAG_CRITERIA['POOR_PERFORMANCE']['minSales'] and
            row['Clicks'] < self.TAG_CRITERIA['POOR_PERFORMANCE']['minClicks'] and
                (row['ACOS'] == 0 or row['ACOS'] > self.TAG_CRITERIA['POOR_PERFORMANCE']['maxAcos'])):
            tags.append(self.TAG_CRITERIA['POOR_PERFORMANCE']['label'])
        return tags

    def export_to_excel(self, campaign_data, output_path):
        """將分析結果輸出到Excel"""
        # 計算總體指標
        total_metrics = {
            '總廣告活動數': len(campaign_data),
            '有效廣告活動數': len(campaign_data[~campaign_data['Tags'].apply(lambda x: '低成效' in x)]),
            '總銷售額': campaign_data['Sales'].sum(),
            '總支出': campaign_data['Spend'].sum(),
            '總訂單數': campaign_data['7 Day Total Orders (#)'].sum(),
            '總曝光數': campaign_data['Impressions'].sum(),
            '總點擊數': campaign_data['Clicks'].sum(),
            '整體ACOS': (campaign_data['Spend'].sum() / campaign_data['Sales'].sum() * 100) if campaign_data['Sales'].sum() > 0 else 0,
            '整體ROAS': (campaign_data['Sales'].sum() / campaign_data['Spend'].sum()) if campaign_data['Spend'].sum() > 0 else 0
        }

        # 創建Excel writer
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # 概覽表
            summary_df = pd.DataFrame({
                '指標': total_metrics.keys(),
                '數值': [
                    f"{total_metrics['總廣告活動數']:,d}",
                    f"{total_metrics['有效廣告活動數']:,d}",
                    f"${total_metrics['總銷售額']:,.2f}",
                    f"${total_metrics['總支出']:,.2f}",
                    f"{total_metrics['總訂單數']:,d}",
                    f"{total_metrics['總曝光數']:,d}",
                    f"{total_metrics['總點擊數']:,d}",
                    f"{total_metrics['整體ACOS']:.2f}%",
                    f"{total_metrics['整體ROAS']:.2f}"
                ]
            })
            summary_df.to_excel(writer, sheet_name='概覽', index=False)

            # 詳細數據表
            detail_df = campaign_data.copy()
            # 重命名欄位
            detail_df = detail_df.rename(columns={
                'Campaign Name': '廣告活動名稱',
                'Sales': '銷售額',
                'Spend': '支出',
                '7 Day Total Orders (#)': '訂單數',
                'Impressions': '曝光數',
                'Clicks': '點擊數',
                'CTR': '點擊率',
                'Conversion_Rate': '轉換率',
                'ROAS': 'ROAS',
                'Tags': '標籤',
                'ACOS': 'ACOS'
            })

            # 排序（按銷售額降序）
            detail_df = detail_df.sort_values('銷售額', ascending=False)

            # 格式化數值
            # detail_df['銷售額'] = detail_df['銷售額'].apply(lambda x: f"${x:,.2f}")
            # detail_df['支出'] = detail_df['支出'].apply(lambda x: f"${x:,.2f}")
            # detail_df['ACOS'] = detail_df['ACOS'].apply(lambda x: f"{x:.2f}%")
            # detail_df['點擊率'] = detail_df['點擊率'].apply(lambda x: f"{x:.2f}%")
            # detail_df['轉換率'] = detail_df['轉換率'].apply(lambda x: f"{x:.2f}%")
            # detail_df['ROAS'] = detail_df['ROAS'].apply(lambda x: f"{x:.2f}")
            detail_df['標籤'] = detail_df['標籤'].apply(lambda x: ', '.join(x))

            # 輸出到Excel
            detail_df.to_excel(writer, sheet_name='詳細數據', index=False)

            # 獲取工作簿和工作表
            workbook = writer.book
            summary_sheet = workbook['概覽']
            detail_sheet = workbook['詳細數據']

            # 設定欄寬
            for sheet in [summary_sheet, detail_sheet]:
                for column in sheet.columns:
                    max_length = 0
                    column = [cell for cell in column]
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    sheet.column_dimensions[get_column_letter(
                        column[0].column)].width = adjusted_width

            # 添加表頭樣式
            header_fill = PatternFill(
                start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(color="FFFFFF", bold=True)

            for sheet in [summary_sheet, detail_sheet]:
                for cell in sheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center')

    def analyze_campaigns(self, file1_path, file2_path, output_path):
        """執行完整的分析流程並輸出到Excel"""
        # 載入數據
        df = self.load_data(file1_path, file2_path)

        # 處理數據
        campaign_metrics = self.process_data(df)

        # 添加標籤
        campaign_metrics['Tags'] = campaign_metrics.apply(
            self.get_campaign_tags, axis=1)

        # 輸出到Excel
        self.export_to_excel(campaign_metrics, output_path)

        return campaign_metrics

    def create_trend_chart(self, df, campaign_name):
        """為單個廣告活動創建時間趨勢圖"""
        campaign_data = df[df['Campaign Name'] == campaign_name].copy()
        campaign_data['Start Date'] = pd.to_datetime(
            campaign_data['Start Date'])
        campaign_data = campaign_data.sort_values('Start Date')

        # 計算每日的 ACOS
        campaign_data['Daily_ACOS'] = np.where(
            campaign_data['Sales'] > 0,
            (campaign_data['Spend'] / campaign_data['Sales'] * 100),
            0
        )

        # 創建圖表
        fig, (ax1, ax2) = plt.subplots(
            2, 1, figsize=(12, 8), height_ratios=[2, 1])
        fig.suptitle(f'廣告活動趨勢分析: {campaign_name}', fontsize=12)

        # 銷售額和支出趨勢
        ax1.plot(campaign_data['Start Date'], campaign_data['Sales'],
                 label='銷售額', color='blue', marker='o')
        ax1.plot(campaign_data['Start Date'], campaign_data['Spend'],
                 label='支出', color='red', marker='s')
        ax1.set_ylabel('金額 ($)')
        ax1.legend()
        ax1.grid(True, linestyle='--', alpha=0.7)

        # ACOS 趨勢
        ax2.plot(campaign_data['Start Date'], campaign_data['Daily_ACOS'],
                 label='ACOS', color='green', marker='o')
        ax2.set_ylabel('ACOS (%)')
        ax2.set_xlabel('日期')
        ax2.grid(True, linestyle='--', alpha=0.7)

        # 格式化日期軸
        for ax in [ax1, ax2]:
            ax.xaxis.set_major_formatter(mdates.DateFormatter('%m/%d'))
            ax.tick_params(axis='x', rotation=45)

        plt.tight_layout()

        # 將圖表轉換為圖片數據
        img_data = BytesIO()
        plt.savefig(img_data, format='png', bbox_inches='tight')
        plt.close()

        return img_data

    def export_to_excel(self, campaign_data, df_raw, output_path):
        """擴展的 Excel 輸出方法，增加績優廣告活動分析"""
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # ... (原有的概覽和詳細數據表保持不變)

            # 新增 ACOS 優異廣告活動分析
            good_acos_campaigns = campaign_data[
                campaign_data['Tags'].apply(
                    lambda x: self.TAG_CRITERIA['LOW_ACOS']['label'] in x)
            ].copy()

            if not good_acos_campaigns.empty:
                # 創建新的工作表
                workbook = writer.book
                good_acos_sheet = workbook.create_sheet('ACOS優異廣告活動')

                # 為每個優異廣告活動創建分析圖表
                row_position = 1
                for idx, campaign in good_acos_campaigns.iterrows():
                    # 添加廣告活動基本信息
                    good_acos_sheet.cell(
                        row=row_position, column=1, value=campaign['Campaign Name'])
                    good_acos_sheet.cell(
                        row=row_position+1, column=1, value='總銷售額')
                    good_acos_sheet.cell(
                        row=row_position+1, column=2, value=f"${campaign['Sales']:,.2f}")
                    good_acos_sheet.cell(
                        row=row_position+2, column=1, value='總支出')
                    good_acos_sheet.cell(
                        row=row_position+2, column=2, value=f"${campaign['Spend']:,.2f}")
                    good_acos_sheet.cell(
                        row=row_position+3, column=1, value='ACOS')
                    good_acos_sheet.cell(
                        row=row_position+3, column=2, value=f"{campaign['ACOS']:.2f}%")

                    # 創建並插入趨勢圖
                    img_data = self.create_trend_chart(
                        df_raw, campaign['Campaign Name'])
                    img = openpyxl.drawing.image.Image(img_data)
                    img.anchor = f'D{row_position}'
                    good_acos_sheet.add_image(img)

                    row_position += 25  # 為下一個廣告活動預留空間

                # 調整欄寬
                for column in good_acos_sheet.columns:
                    max_length = 0
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    good_acos_sheet.column_dimensions[get_column_letter(
                        column[0].column)].width = adjusted_width

    def analyze_campaigns(self, file1_path, file2_path, output_path):
        """修改後的分析流程"""
        # 載入數據
        df = self.load_data(file1_path, file2_path)

        # 處理數據
        campaign_metrics = self.process_data(df)

        # 添加標籤
        campaign_metrics['Tags'] = campaign_metrics.apply(
            self.get_campaign_tags, axis=1)

        # 輸出到Excel（傳入原始數據用於生成趨勢圖）
        self.export_to_excel(campaign_metrics, df, output_path)

        return campaign_metrics


if __name__ == "__main__":
    analyzer = CampaignAnalyzer()

    # 檔案路徑
    file1_path = "./data/241216_Sponsored Products Campaign report.csv"
    file2_path = "./data/241223_Sponsored Products Campaign report.csv"
    output_path = f"廣告活動分析_{datetime.now().strftime('%Y%m%d')}.xlsx"

    # 執行分析並輸出
    campaign_data = analyzer.analyze_campaigns(
        file1_path, file2_path, output_path)
    print(f"分析完成，結果已輸出到: {output_path}")
